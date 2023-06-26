import openpyxl as xl


def write_xls(lst: list, path: str):
    wb = xl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save(path)


def excel_reader(file_obj: str, pagename: str, column: str):
    result = []
    wb = xl.load_workbook(file_obj)
    sheet = wb[pagename]
    for row in sheet[column]:
        result.append(str(row.value))
    return result
